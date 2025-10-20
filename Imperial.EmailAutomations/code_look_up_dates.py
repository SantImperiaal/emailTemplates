# Install required packages:
# pip install selenium openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
import pyperclip
import sys

# Set up Edge WebDriver (download msedgedriver if needed)
driver = webdriver.Edge()  # Make sure msedgedriver.exe is in your PATH

# 1. Open browser and go to URL
driver.get("https://appnav-prod.ec.imperial.ac.uk/applicationNavigator/seamless")

# First page uses a constant code (set the value below)
FIRST_PAGE_CODE = "SZISTUR"  # <-- put your constant here (leave as "" to prompt instead)

if FIRST_PAGE_CODE:
    first_code = FIRST_PAGE_CODE
else:
    first_code = input("Enter first-page code: ").strip()
    if not first_code:
        print("No code provided. Exiting.")
        sys.exit(1)

# wait for the first input and submit the constant/prompted code (using XPath)
first_input = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div/div/main/center-search/div/div/div/div[2]/div[2]/div/input[1]"))
)
first_input.clear()
first_input.send_keys(first_code.upper())
first_input.send_keys(Keys.RETURN)

# --- inserted: prompt for the ID and paste into the search box ---
id_value = input("Enter ID to search (e.g. TFDE12345): ").strip()
if not id_value:
    print("No ID provided. Exiting.")
    sys.exit(1)

# wait for the search input to appear (use the element id you provided)
search_input = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, "inp:key_block_id"))
)

# Set the value via JS (reliable paste), dispatch input event, then submit
driver.execute_script(
    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
    search_input,
    id_value.upper()
)
search_input.send_keys(Keys.RETURN)
# --- end inserted ---

# Now the second page/input differs — prompt user for it (optional)
second_code = input("If a second code is required, enter it now (or press Enter to skip): ").strip()
if second_code:
    try:
        next_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "inp:second_code"))  # replace with real id if different
        )
        next_input.clear()
        next_input.send_keys(second_code)
        next_input.send_keys(Keys.RETURN)
    except Exception as e:
        print("Second input not found or timed out:", e)

# wait for results to load before extracting dates
date1_elem = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "date1")))
date2_elem = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "date2")))
date1 = date1_elem.text
date2 = date2_elem.text

# 4. Write both dates into a single cell (column B). If the ID exists in column A update that row.
excel_path = "C:\\repos\\emailTemplates\\CodeLookUp\\Instalments Code Lookup.xlsx"
combined_dates = f"{date1} | {date2}"

try:
    wb = load_workbook(excel_path)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Dates"])

# Search for existing ID in column A
row_found = None
for cell in ws["A"]:
    if cell.value == id_value:
        row_found = cell.row
        break

if row_found:
    ws.cell(row=row_found, column=2, value=combined_dates)
else:
    ws.append([id_value, combined_dates])

wb.save(excel_path)

# Copy ID to clipboard
pyperclip.copy(id_value)
print("ID copied to clipboard. Paste it (Ctrl+V) where needed.")

driver.quit()